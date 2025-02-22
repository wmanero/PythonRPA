'''''''''4'''# -*- coding: utf-8 -*-
"""
Created on Wed Oct 30 20:11:11 2024

@author: Welton Silva
"""

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from openpyxl import Workbook, load_workbook
import time

#Invoca o método By
from selenium.webdriver.common.by import By

def inserir_dados_arq(arquivo, dados):
    # Nome arquivo e dados        
    try:
        # Carregar arquivo existente ou cria novo
        try:
            wb = load_workbook(arquivo)
            ws = wb.active
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
        
        # Inserir dados na próxima linha vazia
        for linha in dados:
            ws.append(linha)
            
        # Salvar arquivo
        wb.save(arquivo)
        print(f'Dados salvos com sucesso em "{arquivo}"')
        
    except Exception as e:
        print(f'Falha ao tentar salvar dados: {e}')
     

def pesquisa(s):   
    
    dados = []
    
    s_dir = Service(r"C:\\webdriver\\chromedriver.exe")
    driver = webdriver.Chrome(service=s_dir)
    driver.get("https://concla.ibge.gov.br/busca-online-cnae.html?view=estrutura")
    driver.maximize_window()
    
    #Rola tela até seção
    driver.find_element(By.XPATH, '//*[@id="tbEstrutura"]/thead/tr/th[1]').location_once_scrolled_into_view
    time.sleep(2)
    
    # cod seção
    cod_secao = driver.find_element(By.XPATH, f'//*[@id="tbEstrutura"]/tbody/tr[{s}]/td[1]/a').text
    print(cod_secao)
    
    #clica no link da seção
    driver.find_element(By.XPATH, f'//*[@id="tbEstrutura"]/tbody/tr[{s}]/td[1]/a').click()
    time.sleep(2)

    #rola tela até divisão 01
    driver.find_element(By.XPATH, '//*[@id="hierarquia"]/table/tbody/tr[1]/td[1]/span').location_once_scrolled_into_view
    time.sleep(2)
    
    # nome da divisão 01
    nome_divisao_01 = driver.find_element(By.XPATH, '//*[@id="hierarquia"]/table/tbody/tr[1]/td[3]/span').text
    print(nome_divisao_01)
    # cod divisão 01
    cod_divisao = driver.find_element(By.XPATH, '//*[@id="hierarquia"]/table/tbody/tr[2]/td[3]/a').text
 
    # nome da divisão 02
    nome_divisao_02 = driver.find_element(By.XPATH, '//*[@id="hierarquia"]/table/tbody/tr[2]/td[3]').text  
    print(nome_divisao_02)
    #clica na divisão 01
    driver.find_element(By.XPATH, '//*[@id="hierarquia"]/table/tbody/tr[2]/td[3]/a').click()
    time.sleep(2)
    
    #rolar tela até seção
    driver.find_element(By.XPATH, '//*[@id="hierarquia"]/table/tbody/tr[1]/td[1]/span').location_once_scrolled_into_view
    time.sleep(2)
    
    # cod do grupo
    cod_grupo = driver.find_element(By.XPATH, '//*[@id="hierarquia"]/table/tbody/tr[3]/td[3]/a').text
    print(cod_grupo)
    # nome do grupo
    nome_grupo = driver.find_element(By.XPATH, '//*[@id="hierarquia"]/table/tbody/tr[3]/td[3]').text
    print(nome_grupo)
    #clica no grupo
    driver.find_element(By.XPATH, '//*[@id="hierarquia"]/table/tbody/tr[3]/td[3]/a').click()
    time.sleep(2)
    
    #rola tela seção
    driver.find_element(By.XPATH, '//*[@id="hierarquia"]/table/tbody/tr[1]/td[1]/span').location_once_scrolled_into_view
    time.sleep(2)
    
    # cod classe
    cod_classe = driver.find_element(By.XPATH, '//*[@id="hierarquia"]/table/tbody/tr[4]/td[3]/a').text
    print(cod_classe)
    # nome classe
    nome_classe = driver.find_element(By.XPATH, '//*[@id="hierarquia"]/table/tbody/tr[4]/td[3]').text
    print(nome_classe)
    #clica no cod da classe
    driver.find_element(By.XPATH, '//*[@id="hierarquia"]/table/tbody/tr[4]/td[3]/a').click()
    time.sleep(2)
    
    # subclasse
    
    #rola tela até classe
    driver.find_element(By.XPATH, '//*[@id="hierarquia"]/table/tbody/tr[5]/td[1]/span').location_once_scrolled_into_view
    time.sleep(2)
 
    dados = [cod_secao, cod_divisao, nome_divisao_01, nome_divisao_02, cod_grupo, nome_grupo, cod_classe, nome_classe]  
 
    try:
        if s == 1:
            x = 9
        elif s == 2:
            x = 7
        else:
            x = 10
            
        for i in range(5, x):
            # cod subclasse 1
            cod_subclasse_01 = driver.find_element(By.XPATH, f'//*[@id="hierarquia"]/table/tbody/tr[{i}]/td[3]/a').text
            print(cod_subclasse_01)
            # nome subclasse 1
            nome_subclasse_01 = driver.find_element(By.XPATH, f'//*[@id="hierarquia"]/table/tbody/tr[{i}]/td[3]').text
            print(nome_subclasse_01)
            
            dados.append[cod_subclasse_01, nome_subclasse_01]

    except Exception as e:
        print(f'Ocorreu o erro: {e}')
    
    finally:
        print("Consulta finalizada.")
        print(dados)
    
    
        #dados = [cod_secao, cod_divisao, nome_divisao_01, nome_divisao_02, cod_grupo, nome_grupo, cod_classe, nome_classe]#, cod_subclasse_01, nome_subclasse_01]
        
        #print(dados)
        #inserir_dados_arq('relatorio.xlsx', dados)  
        
        #dados.clear()
   
        driver.quit()

def todos():
    for i in range(1,4):
        print(i)
        pesquisa(i)

def secao():
    
    print("Escolha a seção:") 
    print("(1) - Para Seção 'A'")
    print("(2) - Para Seção 'B'")
    print("(3) - Para Seção 'C'")
    print("(4) - Para todas as seções")
    print("(5) - Para sair")
    
    try:
        opcao = int(input("Opção: "))
    except:
        print("Opção inválida! Escolha novamente.")
        secao()
    
    op = [ i for i in range(1, 5)]
    
    if opcao == 4:
        todos()
        
    elif opcao == 5:
        print("Até mais...")
         
    elif opcao < 4:
        pesquisa(opcao)
    else:
        print("Por favor, escolha uma opção existente!")
        secao()
          
        
secao()    



