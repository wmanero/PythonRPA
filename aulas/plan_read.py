import os
from openpyxl import load_workbook
"""
print(os.listdir())
"""
wb = load_workbook(".\\aulas\\vendas.xlsx")
wb.sheetnames

sheet = wb['Planilha1']

"""
valor = sheet['A2'].value
linhas = sheet.max_row
print(valor , linhas)
"""

for linha in range(2, sheet.max_row+1):
    produto = sheet.cell(linha, 1).value
    marca = sheet.cell(linha, 2).value
    valor = sheet.cell(linha, 3).value

    print(f"O produto {produto} da marca {marca} custa {valor} reais.")

wb.close()
